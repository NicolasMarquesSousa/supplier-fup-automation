"""Gera relatórios individuais e rascunhos Outlook a partir de bases locais."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import tempfile
import unicodedata
from datetime import date, datetime
from pathlib import Path

import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def normalizar(valor: object) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9]+", " ", texto.encode("ascii", "ignore").decode().upper())).strip()


def nome_seguro(valor: object) -> str:
    return re.sub(r"[<>:\"/\\|?*]+", "-", str(valor or "").strip())


def extrair_emails(valor: object) -> list[str]:
    return list(dict.fromkeys(re.findall(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", str(valor or "").lower())))


def formatar_mes(valor: object) -> str:
    if isinstance(valor, (datetime, date)):
        return valor.strftime("%m/%Y")
    return str(valor or "")


def formatar_planilha(ws) -> None:
    borda = Border(*(Side(style="thin", color="888888") for _ in range(4)))
    for celula in ws[1]:
        celula.font = Font(name="Arial", size=11, bold=True)
        celula.fill = PatternFill("solid", fgColor="E7E6E6")
        celula.border = borda
    for linha in ws.iter_rows(min_row=2):
        for celula in linha:
            celula.font = Font(name="Arial", size=11)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")


def carregar_contatos(caminho: Path) -> dict[tuple[str, str], list[str]]:
    contatos: dict[tuple[str, str], list[str]] = {}
    with caminho.open(encoding="utf-8-sig", newline="") as arquivo:
        for linha in csv.DictReader(arquivo, delimiter=";"):
            chave = (normalizar(linha["Fornecedor"]), normalizar(linha["Unidade"]))
            contatos[chave] = extrair_emails(linha["Email"])
    return contatos


def gerar_relatorio(base: Path, fornecedor: str, unidade: str, destino: Path) -> int:
    wb = load_workbook(base)
    ws = wb["Base Fup"]
    cabecalhos = [celula.value for celula in ws[1]]
    col_fornecedor = cabecalhos.index("Fornecedor") + 1
    col_unidade = cabecalhos.index("Local") + 1
    for linha in range(ws.max_row, 1, -1):
        if normalizar(ws.cell(linha, col_fornecedor).value) != normalizar(fornecedor) or normalizar(ws.cell(linha, col_unidade).value) != normalizar(unidade):
            ws.delete_rows(linha)
    for coluna in ("Função de revisor", "Agente"):
        cabecalhos = [celula.value for celula in ws[1]]
        if coluna in cabecalhos:
            ws.delete_cols(cabecalhos.index(coluna) + 1)
    if "Data inicial da SOW" in [celula.value for celula in ws[1]]:
        indice = [celula.value for celula in ws[1]].index("Data inicial da SOW") + 1
        for linha in range(2, ws.max_row + 1):
            ws.cell(linha, indice).value = formatar_mes(ws.cell(linha, indice).value)
    if "Planilha1" in wb.sheetnames:
        del wb["Planilha1"]
    formatar_planilha(ws)
    wb.save(destino)
    wb.close()
    return ws.max_row - 1


def executar(base: Path, contatos_csv: Path, saida: Path, enviar: bool = False) -> None:
    saida.mkdir(parents=True, exist_ok=True)
    contatos = carregar_contatos(contatos_csv)
    wb = load_workbook(base, read_only=True, data_only=True)
    ws = wb["Base Fup"]
    headers = [celula.value for celula in ws[1]]
    cf, cl = headers.index("Fornecedor") + 1, headers.index("Local") + 1
    grupos = sorted({(str(ws.cell(i, cf).value).strip(), str(ws.cell(i, cl).value or "SEM LOCAL").strip()) for i in range(2, ws.max_row + 1) if ws.cell(i, cf).value})
    wb.close()
    outlook = win32.Dispatch("Outlook.Application")
    for fornecedor, unidade in grupos:
        anexo = saida / f"FUP - {nome_seguro(fornecedor)} - {nome_seguro(unidade)}.xlsx"
        pendencias = gerar_relatorio(base, fornecedor, unidade, anexo)
        email = outlook.CreateItem(0)
        email.To = ";".join(contatos.get((normalizar(fornecedor), normalizar(unidade)), []))
        email.Subject = f"Follow-up de documentação pendente — {fornecedor}"
        email.HTMLBody = f"<p>Olá, equipe do fornecedor.</p><p>O relatório anexo contém {pendencias} pendência(s). Revise os dados e os prazos.</p>"
        copia = Path(tempfile.gettempdir()) / anexo.name
        shutil.copy2(anexo, copia)
        email.Attachments.Add(str(copia))
        if enviar and email.To:
            email.Send()
        else:
            email.Save()
        copia.unlink(missing_ok=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", type=Path, required=True)
    parser.add_argument("--contatos", type=Path, required=True)
    parser.add_argument("--saida", type=Path, default=Path("output"))
    parser.add_argument("--enviar", action="store_true", help="Envia mensagens; sem esta opção, apenas cria rascunhos")
    args = parser.parse_args()
    executar(args.base, args.contatos, args.saida, args.enviar)
